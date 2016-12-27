#!/usr/bin/env python
import math
import sys
import json
import os
import datetime
import time
import re
import datetime
import collections
import openpyxl
import unicodedata
from Debtor import Debtor
from sets import Set
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from xlrd import open_workbook
from optparse import OptionParser

class ReadExcel:

	def __init__(self):
		# The list of files
		self.files = []

	def parseFile(self, excelFilename):
		jsonList    = []
		workbook    = open_workbook(excelFilename)
		sheet_names = workbook.sheet_names()
		sheet       = workbook.sheet_by_index(0)
		header      = None
		self.files.append(excelFilename)

		for row in range(sheet.nrows):
			values = []
			for col in range(sheet.ncols):
				cellValue   = sheet.cell(row,col).value
				cleanedText = self.cleanText(cellValue)
				values.append(cleanedText)

			if values[0].lower() == 'name':
				header = self.getHeader(values)
			elif values[1] != '':
				jsonValues = self.asJSON(header, values)
				jsonList.append(jsonValues)
		return jsonList

	def asJSON(self, header, valueList):
		jsonDict = collections.OrderedDict()
		errors   = 0

		if header is not None:
			nameIndex   = header.index('name')
			addrIndex   = header.index('address')
			occIndex    = header.index('occupation')
			jsonDict['name']       = valueList[nameIndex]
			jsonDict['occupation'] = valueList[occIndex]

			if 'county' in header:
				countyIndex = header.index('county')
				jsonDict['address']    = valueList[addrIndex] + ' ' + valueList[countyIndex]
			else:
				jsonDict['address']    = valueList[addrIndex]
		else:
			jsonDict['name']       = valueList[0]
			jsonDict['address']    = valueList[1]
			jsonDict['occupation'] = valueList[2]
			
		jsonStr = json.dumps(jsonDict, ensure_ascii=True)
		return jsonStr

	def getHeader(self, valueList):
		outputList = [x.lower() for x in valueList]
		return outputList

	def printDebtors(self):
		for debtor in self.debtorList:
			debtor.printDetails()

	def isHeader(self, valueList):
		if len(valueList) > 4 and valueList[0].lower() != 'name' and valueList[1] != '':
			return True
		return False

	def cleanText(self, inputText):
		if inputText is None:
			return ''
		elif type(inputText) is unicode:
			stringValue = unicodedata.normalize('NFKD', inputText).encode('ascii','ignore')
			stringValue = stringValue.replace('\n', ' ')
			stringValue = re.sub( '\s+', ' ', stringValue).strip()
			return stringValue
		else:
			stringValue = str(inputText)
			stringValue = stringValue.replace('\n', ' ')
			stringValue = re.sub( '\s+', ' ', stringValue).strip()
			return stringValue


	def processXLSX(self, excelFilename):

		# Read in the Excel file workbook
		workbook  = openpyxl.load_workbook(excelFilename, data_only=True)

		# Only evaluate the top worksheet
		worksheet = workbook.worksheets[0]

		rowLimit = worksheet.max_row
		for x in xrange(1, rowLimit):
			if worksheet.cell(row=x, column=1).value is not None:
				data1 = worksheet.cell(row=x, column=1).value
				print data1

def main(argv):
	parser = OptionParser(usage="Usage: ReadExcelFile <excel-filename>")

        parser.add_option("-v", "--verbose",
                action="store_true",
                dest="verboseFlag",
                default=False,
                help="Verbose output from the script")

	(options, filename) = parser.parse_args()

	if len(filename) != 1 or not os.path.isfile(filename[0]) :
		print parser.print_help()
		exit(1)

	check = ReadExcel()
	jsonString = check.parseFile(filename[0])
	print jsonString
		
if __name__ == "__main__":
    sys.exit(main(sys.argv))
